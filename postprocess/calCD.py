"""
计算各种方法与groundtruth之间的CD值，具体计算方法是：读取IBS面gt，读取其他算法的点云形式IBS面，计算点云到IBS面的距离，求均值
"""
import open3d as o3d
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from utils.path_utils import *


def parseConfig(config_filepath: str = './config/visualization.json'):
    with open(config_filepath, 'r') as configfile:
        specs = json.load(configfile)

    return specs


def getGeometriesPath(specs, filename):
    category_re = specs["category_re"]
    scene_re = specs["scene_re"]
    filename_re = specs["filename_re"]
    category = re.match(category_re, filename).group()
    scene = re.match(scene_re, filename).group()
    filename = re.match(filename_re, filename).group()

    geometries_path = dict()

    ibs_gt_mesh_dir = specs['ibs_gt_mesh_dir']
    ibs_pred_pcd_dir = specs['ibs_pred_pcd_dir']

    ibs_gt_mesh_filename = '{}.obj'.format(scene)
    ibs_pred_pcd_filename = '{}.ply'.format(filename)

    geometries_path['ibs_gt_mesh'] = os.path.join(ibs_gt_mesh_dir, category, ibs_gt_mesh_filename)
    geometries_path['ibs_pred_pcd'] = os.path.join(ibs_pred_pcd_dir, category, ibs_pred_pcd_filename)

    return geometries_path


class Comparator:
    def __init__(self, specs):
        self.cd = 0
        self.specs = specs

    def get_pcd(self, pcd_filepath):
        pcd = o3d.io.read_point_cloud(pcd_filepath)
        return pcd

    def get_mesh(self, mesh_filepath):
        mesh = o3d.io.read_triangle_mesh(mesh_filepath)
        return mesh

    def query_dist(self, mesh, points):
        points = o3d.core.Tensor(np.array(points), dtype=o3d.core.Dtype.Float32)
        scene = o3d.t.geometry.RaycastingScene()
        mesh_t = o3d.t.geometry.TriangleMesh.from_legacy(mesh)
        scene.add_triangles(mesh_t)
        return scene.compute_distance(points)

    def calCD(self, ibs_gt_mesh, ibs_pred_pcd):
        self.cd = self.query_dist(ibs_gt_mesh, ibs_pred_pcd.points).numpy()
        self.cd = sum(self.cd) / len(self.cd)

    def handle_current_file(self, specs, filename):
        geometries_path = getGeometriesPath(specs, filename)
        ibs_gt_mesh = self.get_mesh(geometries_path['ibs_gt_mesh'])
        ibs_pred_pcd = self.get_pcd(geometries_path['ibs_pred_pcd'])

        self.calCD(ibs_gt_mesh, ibs_pred_pcd)


class categoryTableGenerator:
    def __init__(self, wb, category):
        self.ws = wb.create_sheet(category)
        self.ws.cell(row=1, column=1, value="scene_no")
        self.ws.cell(row=2, column=1, value="cd")
        self.avg_cd = 0

    def convert_data(self, cd_dict):
        name_list = []
        cd_list = []
        for scene in cd_dict:
            name_list.append(scene)
            cd_cur_scene = []
            for filename in cd_dict[scene]:
                cd_cur_scene.append(cd_dict[scene][filename])
            cd_list.append(sum(cd_cur_scene)/len(cd_cur_scene))
        return name_list, cd_list

    def save_result(self, cd_dict):
        name_list, cd_list = self.convert_data(cd_dict)
        for i, scene_name in enumerate(name_list):
            self.ws.cell(row=1, column=i+2, value=name_list[i])
            self.ws.cell(row=2, column=i+2, value=cd_list[i])
        self.avg_cd = sum(cd_list)/len(cd_list)
        self.ws.cell(row=1, column=len(name_list)+2, value="avg")
        self.ws.cell(row=2, column=len(name_list)+2, value=self.avg_cd)
        for i in range(1, len(name_list)+3):
            self.ws.column_dimensions[get_column_letter(i)].width = 15


def save_category_cd(wb, category_cd_dict):
    ws = wb.create_sheet("total")
    ws.cell(row=1, column=1, value="category")
    ws.cell(row=2, column=1, value="cd")

    name_list = []
    cd_list = []
    for category in category_cd_dict:
        name_list.append(category)
        cd_list.append(category_cd_dict[category])

    for i, scene_name in enumerate(name_list):
        ws.cell(row=1, column=i + 2, value=name_list[i])
        ws.cell(row=2, column=i + 2, value=cd_list[i])
    avg_cd = sum(cd_list) / len(cd_list)
    ws.cell(row=1, column=len(name_list) + 2, value="avg")
    ws.cell(row=2, column=len(name_list) + 2, value=avg_cd)
    for i in range(1, len(name_list)+3):
        ws.column_dimensions[get_column_letter(i)].width = 15


if __name__ == '__main__':
    # 获取配置参数
    config_filepath = 'config/calCD.json'
    specs = parseConfig(config_filepath)

    filename_tree = getFilenameTree(specs, "ibs_pred_pcd_dir")
    wb = Workbook()
    comparator = Comparator(specs)

    cd_dict = dict()
    category_cd_dict = dict()
    for category in filename_tree:
        cd_dict[category] = dict()
        for scene in filename_tree[category]:
            cd_dict[category][scene] = dict()

    for category in filename_tree:
        generator = categoryTableGenerator(wb, category)
        # 统计当前类别的cd
        scene_name_list = []
        for scene in filename_tree[category]:
            scene_name_list.append(scene)
            # 统计当前场景的cd
            print('current scene: ', scene)
            for filename in filename_tree[category][scene]:
                print('current filename: ', filename)
                comparator.handle_current_file(specs, filename)
                cd_dict[category][scene][filename] = comparator.cd

        generator.save_result(cd_dict[category])
        category_cd_dict[category] = generator.avg_cd

    save_category_cd(wb, category_cd_dict)
    wb.save('{}.xlsx'.format(specs["table_name"]))
    wb.close()

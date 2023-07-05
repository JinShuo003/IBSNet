"""
可视化工具，配置好./config/visualization.json后可以可视化mesh模型、点云、位于模型表面和IBS表面的sdf点、各自和总体的aabb框、交互区域gt
"""
import open3d as o3d
import os
import re
import json
import numpy as np
from openpyxl import Workbook


def parseConfig(config_filepath: str = './config/visualization.json'):
    with open(config_filepath, 'r') as configfile:
        specs = json.load(configfile)

    return specs


def getFilenameTree(specs: dict):
    # 以SDF network作为基准构建文件树
    sdf_path = specs["sdf_indirect_dir"]
    category_re = specs["category_re"]
    scene_re = specs["scene_re"]
    filename_re = specs["filename_re"]

    filename_tree = dict()
    folder_info = os.walk(sdf_path)
    for dir_path, dir_names, filenames in folder_info:
        # 当前是顶级目录，不做处理
        if dir_path == sdf_path:
            continue
        # 获取当前文件夹的类目信息，合法则继续处理
        category = dir_path.split('\\')[-1]
        if not re.match(category_re, category):
            continue
        # 当前类目不在filename_tree中，添加
        if not category in filename_tree:
            filename_tree[category] = dict()
        for filename in filenames:
            if not re.match(filename_re, filename):
                continue
            # 获取场景名
            scene = re.match(scene_re, filename)
            # 如果与场景re不匹配则跳过
            if not scene:
                continue
            scene = scene.group()
            # 当前场景不在filename_tree[category]中，添加
            if not scene in filename_tree[category]:
                filename_tree[category][scene] = list()
            filename_tree[category][scene].append(filename)

    return filename_tree


def getGeometriesPath(specs, filename):
    category_re = specs["category_re"]
    filename_re = specs["filename_re"]
    category = re.match(category_re, filename).group()
    filename = re.match(filename_re, filename).group()

    geometries_path = dict()

    pcd_dir = specs['pcd_dir']
    sdf_scan_dir = specs['sdf_scan_dir']
    sdf_gt_dir = specs['sdf_gt_dir']
    sdf_indirect_dir = specs['sdf_indirect_dir']
    sdf_direct_dir = specs['sdf_direct_dir']

    pcd1_filename = '{}_{}.ply'.format(filename, 0)
    pcd2_filename = '{}_{}.ply'.format(filename, 1)
    sdf_filename = '{}.npz'.format(filename)

    geometries_path['pcd1'] = os.path.join(pcd_dir, category, pcd1_filename)
    geometries_path['pcd2'] = os.path.join(pcd_dir, category, pcd2_filename)
    geometries_path['sdf_scan'] = os.path.join(sdf_scan_dir, category, sdf_filename)
    geometries_path['sdf_gt'] = os.path.join(sdf_gt_dir, category, sdf_filename)
    geometries_path['sdf_indirect_dir'] = os.path.join(sdf_indirect_dir, category, sdf_filename)
    geometries_path['sdf_direct_dir'] = os.path.join(sdf_direct_dir, category, sdf_filename)

    return geometries_path


def get_pcd(specs, pcd_filepath, color_key):
    pcd = o3d.io.read_point_cloud(pcd_filepath)
    pcd.paint_uniform_color(specs['visualization_options']['colors'][color_key])

    return pcd


def get_surface_points(specs, sdf_filepath):
    npz = np.load(sdf_filepath)

    data = npz["data"]
    if data.shape[1] == 5:
        surface_points_ibs = [points[0:3] for points in data if
                              abs(points[3] - points[4]) < specs['visualization_options']['sdf_threshold']]
    elif data.shape[1] == 4:
        surface_points_ibs = [points[0:3] for points in data if
                              points[3] < specs['visualization_options']['sdf_threshold']]
    pcd_ibs = o3d.geometry.PointCloud()
    pcd_ibs.points = o3d.utility.Vector3dVector(surface_points_ibs)

    pcd_ibs.paint_uniform_color(specs['visualization_options']['colors']['ibs'])

    return pcd_ibs


class Scene_visualizer:
    def __init__(self):
        self.vis = None
        self.other_visualizers = []
        self.geometries = []

    def add_other_visualizers(self, visualizer_list):
        for visualizer in visualizer_list:
            if self.vis != visualizer:
                self.other_visualizers.append(visualizer)

    def create_window(self, window_name, width, height, left, top):
        self.vis = o3d.visualization.Visualizer()
        self.vis.create_window(window_name=window_name, width=width, height=height, left=left, top=top)

    def destroy_window(self):
        self.vis.destroy_window()
        self.vis = None

    def add_geometries(self, geometries_list):
        self.geometries = geometries_list
        for geometry in geometries_list:
            self.vis.add_geometry(geometry)

    def update_geometries(self):
        for geometry in self.geometries:
            self.vis.update_geometry(geometry)

    def get_view(self):
        return self.vis.get_view_control().convert_to_pinhole_camera_parameters()

    def set_view(self, view):
        return self.vis.get_view_control().convert_from_pinhole_camera_parameters(view)

    def notify_all(self):
        for visualizer in self.other_visualizers:
            visualizer.set_view(self.get_view())


class Comparator:
    def __init__(self, specs):
        self.chamfer_distance_scan = 0
        self.chamfer_distance_indirect = 0
        self.chamfer_distance_direct = 0
        self.specs = specs
        self.is_cal_CD = specs["cal_CD"]
        self.is_visualize = specs["visualize"]
        self.scene_visualizer_list = [Scene_visualizer(), Scene_visualizer(), Scene_visualizer(), Scene_visualizer()]
        self.bind_windows()

    def bind_windows(self):
        for visualizer in self.scene_visualizer_list:
            visualizer.add_other_visualizers(self.scene_visualizer_list)

    def destroy_windows(self):
        for visualizer in self.scene_visualizer_list:
            visualizer.destroy_window()

    def handle_current_file(self, specs, filename):
        geometries_path = getGeometriesPath(specs, filename)
        pcd1 = get_pcd(specs, geometries_path['pcd1'], 'pcd1')
        pcd2 = get_pcd(specs, geometries_path['pcd2'], 'pcd2')
        ibs_scan = get_surface_points(specs, geometries_path['sdf_scan'])
        ibs_gt = get_surface_points(specs, geometries_path['sdf_gt'])
        ibs_indirect = get_surface_points(specs, geometries_path['sdf_indirect_dir'])
        ibs_direct = get_surface_points(specs, geometries_path['sdf_direct_dir'])

        if self.is_cal_CD:
            self.calCD(ibs_scan, ibs_gt, ibs_indirect, ibs_direct)
        if self.is_visualize:
            self.visualize(pcd1, pcd2, ibs_scan, ibs_gt, ibs_indirect, ibs_direct)

    def calCD(self, ibs_scan, ibs_gt, ibs_indirect, ibs_direct):
        distances_scan = np.asarray(ibs_scan.compute_point_cloud_distance(ibs_gt))
        distances_indirect = np.asarray(ibs_indirect.compute_point_cloud_distance(ibs_gt))
        distances_direct = np.asarray(ibs_direct.compute_point_cloud_distance(ibs_gt))

        self.chamfer_distance_scan = sum(distances_scan) / len(distances_scan)
        self.chamfer_distance_indirect = sum(distances_indirect) / len(distances_indirect)
        try:
            self.chamfer_distance_direct = sum(distances_direct) / len(distances_direct)
        except:
            self.chamfer_distance_direct = -10000

    def visualize(self, pcd1, pcd2, ibs_scan, ibs_gt, ibs_indirect, ibs_direct):
        self.scene_visualizer_list[0].create_window(window_name='传统方法', width=800, height=450, left=0, top=30)
        self.scene_visualizer_list[0].add_geometries([pcd1, pcd2, ibs_scan])
        self.scene_visualizer_list[1].create_window(window_name='groundTruth', width=800, height=450, left=800, top=30)
        self.scene_visualizer_list[1].add_geometries([pcd1, pcd2, ibs_gt])
        self.scene_visualizer_list[2].create_window(window_name='间接法网络预测', width=800, height=450, left=0, top=510)
        self.scene_visualizer_list[2].add_geometries([pcd1, pcd2, ibs_indirect])
        self.scene_visualizer_list[3].create_window(window_name='直接法网络预测', width=800, height=450, left=800, top=510)
        self.scene_visualizer_list[3].add_geometries([pcd1, pcd2, ibs_direct])

        while True:
            self.scene_visualizer_list[0].update_geometries()
            if not self.scene_visualizer_list[0].vis.poll_events():
                break
            self.scene_visualizer_list[0].vis.update_renderer()
            self.scene_visualizer_list[0].notify_all()

            self.scene_visualizer_list[1].update_geometries()
            if not self.scene_visualizer_list[1].vis.poll_events():
                break
            self.scene_visualizer_list[1].vis.update_renderer()
            self.scene_visualizer_list[1].notify_all()

            self.scene_visualizer_list[2].update_geometries()
            if not self.scene_visualizer_list[2].vis.poll_events():
                break
            self.scene_visualizer_list[2].vis.update_renderer()
            self.scene_visualizer_list[2].notify_all()

            self.scene_visualizer_list[3].update_geometries()
            if not self.scene_visualizer_list[3].vis.poll_events():
                break
            self.scene_visualizer_list[3].vis.update_renderer()
            self.scene_visualizer_list[3].notify_all()

        self.destroy_windows()


class categoryTableGenerator:
    def __init__(self, wb, category):
        self.ws = wb.create_sheet(category)
        self.ws.cell(row=1, column=1, value="scene_no")
        self.ws.cell(row=2, column=1, value="partial_scan")
        self.ws.cell(row=3, column=1, value="indirect")
        self.ws.cell(row=4, column=1, value="direct")

    def save_result(self, scene_name_list, chamfer_distance_scan_list, chamfer_distance_indirect_list, chamfer_distance_direct_list):
        for i, scene_name in enumerate(scene_name_list):
            self.ws.cell(row=1, column=i+2, value=scene_name_list[i])
            self.ws.cell(row=2, column=i+2, value=chamfer_distance_scan_list[i])
            self.ws.cell(row=3, column=i+2, value=chamfer_distance_indirect_list[i])
            self.ws.cell(row=4, column=i+2, value=chamfer_distance_direct_list[i])


if __name__ == '__main__':
    # 获取配置参数
    config_filepath = 'config/compare.json'
    specs = parseConfig(config_filepath)

    filename_tree = getFilenameTree(specs)
    wb = Workbook()
    comparator = Comparator(specs)

    for category in filename_tree:
        categoryTableGenerator(wb, category)
        # 统计当前类别的cd
        scene_name_list = []
        chamfer_distance_scan_list = []
        chamfer_distance_indirect_list = []
        chamfer_distance_direct_list = []
        for scene in filename_tree[category]:
            scene_name_list.append(scene)
            # 统计当前场景的cd
            chamfer_distance_scan_scene = 0
            chamfer_distance_indirect_scene = 0
            chamfer_distance_direct_scene = 0
            print('current scene: ', scene)
            for filename in filename_tree[category][scene]:
                comparator.handle_current_file(specs, filename)
                chamfer_distance_scan_scene += comparator.chamfer_distance_scan
                chamfer_distance_indirect_scene += comparator.chamfer_distance_indirect
                chamfer_distance_direct_scene += comparator.chamfer_distance_direct
            chamfer_distance_scan_list.append(chamfer_distance_scan_scene/len(filename_tree[category][scene]))
            chamfer_distance_indirect_list.append(chamfer_distance_indirect_scene/len(filename_tree[category][scene]))
            chamfer_distance_direct_list.append(chamfer_distance_direct_scene/len(filename_tree[category][scene]))
        categoryTableGenerator.save_result(scene_name_list,
                                           chamfer_distance_scan_list,
                                           chamfer_distance_indirect_list,
                                           chamfer_distance_direct_list)

    wb.save('compare_result.xlsx')
    wb.close()
